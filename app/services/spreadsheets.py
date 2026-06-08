import csv
import io
from pathlib import Path

from openpyxl import Workbook, load_workbook


def read_table_file(source: Path | str | io.BytesIO, filename: str = "") -> list[dict[str, str]]:
    # Determine file type from path or filename
    if isinstance(source, (Path, str)) and not isinstance(source, io.BytesIO):
        path = Path(source)
        suffix = path.suffix.lower()
        if suffix == ".csv":
            with path.open("r", encoding="utf-8-sig", newline="") as handle:
                reader = csv.DictReader(handle)
                return [_stringify_row(row) for row in reader]
        if suffix in {".xlsx", ".xlsm"}:
            workbook = load_workbook(path, read_only=True, data_only=True)
        else:
            raise ValueError("仅支持 csv、xlsx、xlsm 文件。")
    else:
        # BytesIO or file-like object
        suffix = Path(filename).suffix.lower() if filename else ".xlsx"
        if suffix == ".csv":
            text = source.read().decode("utf-8-sig") if hasattr(source, "read") else str(source)
            reader = csv.DictReader(io.StringIO(text))
            return [_stringify_row(row) for row in reader]
        if suffix in {".xlsx", ".xlsm"}:
            workbook = load_workbook(source, read_only=True, data_only=True)
        else:
            raise ValueError("仅支持 csv、xlsx、xlsm 文件。")

    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    data: list[dict[str, str]] = []
    for raw_row in rows[1:]:
        row = {headers[index]: _cell_to_text(value) for index, value in enumerate(raw_row) if index < len(headers)}
        if any(value.strip() for value in row.values()):
            data.append(row)
    return data


def write_xlsx(
    file_path: Path,
    headers: list[str],
    rows: list[dict[str, str]],
    summary_rows: list[dict[str, object]] | None = None,
    detail_sheet_name: str = "服务器明细",
    pivot_sheet_name: str = "汇总",
) -> None:
    workbook = Workbook()

    # Sheet 1: Pivot table by 负责人
    pivot_sheet = workbook.active
    pivot_sheet.title = pivot_sheet_name[:31] or "汇总"
    pivot_headers = ["负责人", "服务器ID计数"]
    pivot_sheet.append(pivot_headers)
    for row in summary_rows or [{"负责人": "合计", "服务器ID计数": 0}]:
        pivot_sheet.append([row.get(header, "") for header in pivot_headers])

    # Sheet 2: Server detail
    sheet = workbook.create_sheet(detail_sheet_name[:31] or "服务器明细")
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])
    workbook.save(file_path)


def write_csv(file_path: Path, headers: list[str], rows: list[dict[str, str]]) -> None:
    with file_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def _stringify_row(row: dict[str, object]) -> dict[str, str]:
    return {str(key).strip(): _cell_to_text(value) for key, value in row.items() if key is not None}


def _cell_to_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()
